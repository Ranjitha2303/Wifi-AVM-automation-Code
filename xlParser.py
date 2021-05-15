import openpyxl
import xlrd
import xlrd
import xlwt
from xlwt import Workbook
from PIL import Image
from io import BytesIO
import inspect
import os

import testLogs
import wifiConfig


def readXlFileColums(xlFile = None,
                     rowStart = wifiConfig.xlRowStart,
                     colStart = wifiConfig.xlColStart,
                     numberOfRows = wifiConfig.xlTotalRows):

    if xlFile == None:
        testLogs.ERROR("Please provide XL file to parse")
        exit(1)
    
    testLogs.INFO ("FUNCTION : Reading colum data in XL file : {}".format(xlFile))

    try:
        wb =  openpyxl.load_workbook(xlFile.strip(), read_only = True, data_only=True)
        wb.active = 0
    except Exception as E:
        testLogs.ERROR ("Could not open XL file : {}".format(xlFile))
        testLogs.ERROR (E)

    currentSheet = wb.active
    testLogs.INFO("Current Sheet is : {}".format(currentSheet))
    
    rowEnd = rowStart + numberOfRows
    # Read row values from XL file
    values = []
    for row in range(rowStart, rowEnd):
        testLogs.INFO("Row Col : {} {}".format(row, colStart))
        cellToRead = currentSheet.cell(row = row, column = colStart)
        testLogs.INFO(cellToRead.value)
        values.append(cellToRead.value)

    testLogs.INFO ("Values of Colum No {} is : {}".format(colStart, values))
    return values


def readXlFileRows(xlFile = None,
                   rowStart = wifiConfig.xlRowStart,
                   colStart = wifiConfig.xlColStart,
                   numberOfCols = wifiConfig.xlTotalCols):

    if xlFile == None:
        testLogs.ERROR("Please provide XL file to parse")
        exit(1)
    
    testLogs.INFO ("FUNCTION : Reading row data in XL file : {}".format(xlFile))

    try:
        wb =  openpyxl.load_workbook(xlFile.strip(), data_only=True)
        wb.active = 0
    except Exception as E:
        testLogs.ERROR ("Could not open XL file : {}".format(xlFile))
        testLogs.ERROR (E)

    currentSheet = wb.active

    colEnd = colStart + numberOfCols
    # Read row values from XL file
    values = []
    for col in range(colStart, colEnd):
        testLogs.INFO("Row Col : {} {}".format(rowStart, col))
        cellToRead = currentSheet.cell(row = rowStart, column = col)
        testLogs.INFO(cellToRead.value)
        values.append(cellToRead.value)

    testLogs.INFO ("Values of Row No {} is : {}".format(rowStart, values))
    return values


def writeXlFileColums(xlFile = None,
                      data = [1, 2, 3, 4, 5],
                      rowStart = wifiConfig.xlRowStart,
                      colStart = wifiConfig.xlColStart):

    if xlFile == None:
        testLogs.ERROR("Please provide XL file to parse")
        exit(1)
    
    testLogs.INFO ("FUNCTION : Writing colum data in XL file : {}".format(xlFile))

    fileExist = os.path.exists(xlFile)

    if fileExist != True:
        try:
            wb = openpyxl.Workbook()
            wb.create_sheet("sheet1", 0)
            wb.active = 0
        except Exception as E:
            testLogs.ERROR ("Could not create XL file : {}".format(xlFile))
            testLogs.ERROR (E)
    else:
        wb =  openpyxl.load_workbook(xlFile.strip(), data_only=True)

    currentSheet = wb.active
    

    rowEnd = rowStart + len(data)
    for index, row in enumerate(range(rowStart, rowEnd)):
        cellToModify = currentSheet.cell(row = row, column = colStart)
        cellToModify.value = data[index]

    wb.save(xlFile)


def writeXlFileRows(xlFile = None,
                      data = [1, 2, 3, 4, 5],
                      rowStart = wifiConfig.xlRowStart,
                      colStart = wifiConfig.xlColStart):

    if xlFile == None:
        testLogs.ERROR("Please provide XL file to parse")
        exit(1)
    
    testLogs.INFO ("FUNCTION : Writing colum data in XL file : {}".format(xlFile))

    fileExist = os.path.exists(xlFile)

    if fileExist != True:
        try:
            wb = openpyxl.Workbook()
            wb.create_sheet("sheet1", 0)
            wb.active = 0
        except Exception as E:
            testLogs.ERROR ("Could not create XL file : {}".format(xlFile))
            testLogs.ERROR (E)
    else:
        wb =  openpyxl.load_workbook(xlFile.strip(), data_only=True)

    currentSheet = wb.active
    

    colEnd = colStart + len(data)
    for index, col in enumerate(range(colStart, colEnd)):
        cellToModify = currentSheet.cell(row = rowStart, column = col)
        cellToModify.value = data[index]

    wb.save(xlFile)


def insertXlImage(xlFile = None,
                  xlImage = None,
                  row = 1,
                  col = 1):

    if xlFile == None:
        testLogs.ERROR("Please provide XL file to parse")
        exit(1)   

    if xlImage == None:
        testLogs.ERROR("Please provide Image file to insert")
        exit(1)
    
    testLogs.INFO ("FUNCTION : Inser a Image {} file into XL file : {}".format(xlImage, xlFile))

    fileExist = os.path.exists(xlFile)

    if fileExist != True:
        try:
            wb = openpyxl.Workbook()
            wb.create_sheet("sheet1", 0)
            wb.active = 0
        except Exception as E:
            testLogs.ERROR ("Could not create XL file : {}".format(xlFile))
            testLogs.ERROR (E)
    else:
        testLogs.INFO (xlFile)
        testLogs.INFO (xlFile.strip())
        wb =  openpyxl.load_workbook(xlFile, data_only=True)

    ws = wb.active
    img = openpyxl.drawing.image.Image(xlImage)
    ws.add_image(img, '{}{}'.format(chr(col + 64), row))
    wb.save(xlFile)    


if __name__ == '__main__':
    
    wifiConfig.parseWifiCommandLineArgs()
    #readXlFileRows(wifiConfig.xlToParse,
    #               rowStart = wifiConfig.xlRowStart,
    #               colStart = wifiConfig.xlColStart,
    #               numberOfCols = wifiConfig.xlTotalCols)
    #readXlFileColums(wifiConfig.xlToParse,)
                     #rowStart = wifiConfig.xlRowStart,
                     #colStart = wifiConfig.xlColStart,
                     #numberOfRows = wifiConfig.xlTotalRows)
    #writeXlFileColums(xlFile = "mani.xls")
    insertXlImage(wifiConfig.xlToCreate,
                  b"C:\Users\zilogic\Desktop\TEST_AUTOMATION\AVM_TEST_AUTOMATION\out\graphs\HKv2_MS_5G_TCP_DL_barChart.png",
                  wifiConfig.xlRowStart,
                  wifiConfig.xlColStart)
